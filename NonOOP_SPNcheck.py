import openpyxl as openpy
import numpy as np
import os

# WORKING DIRECTORY MANIPULATION _______________________________________________________________________________________________________________________________________________________________________

working_directory = os.getcwd()
print('')
print("  Working Directory:", working_directory)
os.chdir(r'P:/Active Projects/P9624 777x PSU/04_Core Team/E_Design/Oxygen Integration BE')
ConsolidatedDir = os.getcwd()
print("  New Directory:", ConsolidatedDir)
print('')

# GRAB WORKBOOK, DEFINE WORKSHEETS _____________________________________________________________________________________________________________________________________________________________________

wb = openpy.load_workbook(filename='BODS (WEIGHT & CGs) CONS.xlsx', data_only=True)
wb2 = openpy.load_workbook(filename='BODS (WEIGHT & CGs) CONS DUMP.xlsx')
ws1 = wb['BODS WT']  # Sheet 1 name
ws2 = wb['CG Locations']  # Sheet 2 name
ws3 = wb['Moment Arms']  # Sheet 3 name
ws4 = wb['Dash Numbers']  # Sheet 4 name
ws5 = wb2['RawDataDump']  # Sheet 1 name
# ws6 = wb2['']  # Sheet 3 name

# STYLE GENERATION _____________________________________________________________________________________________________________________________________________________________________________________

'''
MainStyle = NamedStyle(name='MainStyle')
MainStyle.Font(name='Monospac821 BT',
               size=9,
               bold=False)

MainStyle.Alignment(horizontal='general',
                    vertical='general',
                    wrap_text=False,
                    indent=0)

TableStyle.Font(name='Century Gothic',
                size=9,
                bold=False)

TableStyle.Alignment(horizontal='general',
                     vertical='general',
                     wrap_text=False,
                     indent=0)

wb2.add_named_style(MainStyle)
ws6.add_names_style(TableStyle)
'''

# PULL FROM CG MATRIX __________________________________________________________________________________________________________________________________________________________________________________

colcount = ws2.max_column  # Get no. for col quantity
coldiv = int(colcount/5)  # Break columns into groups of 5
rowcount = ws2.max_row  # Get no. for row quantity
title = []
item = []
CGx = []
CGy = []
CGz = []
Mass = []
i = 1
groupiter = 1
group = range(1, coldiv+1)  # Create an array for the overall column groups

# CGs __________________________________________________________________________________________________________________________________________________________________________________________________

while groupiter <= 5:  # Iterate 8 times
    title.append(ws2.cell(row=1, column=int((5*groupiter)-4)).value)  # Create and append title array
    for i in range(1, int((colcount/coldiv)+1)):  # Iterate 5 times
        col = ((groupiter-1)*5) + i  # Move to new column group
    for row in range(3, int(rowcount+1)):  # Go through row for a section, append matrices
        if ws2.cell(row=row, column=col).value is not None:  # Check to ensure there is a value present
            item.append(ws2.cell(row=row, column=int((5*groupiter)-4)).value)
            CGx.append(ws2.cell(row=row, column=int((5*groupiter)-3)).value)
            CGy.append(ws2.cell(row=row, column=int((5*groupiter)-2)).value)
            CGz.append(ws2.cell(row=row, column=int((5*groupiter)-1)).value)
            Mass.append(ws2.cell(row=row, column=int(5*groupiter)).value)
        else:
            continue  # If there is no value, move to new group
    groupiter += 1

print('CG APPEND END!')

i = 1
groupiter = 1
title2 = []
item2 = []
CGx2 = []
CGy2 = []
CGz2 = []
Mass2 = []

# MOMENT ARMS __________________________________________________________________________________________________________________________________________________________________________________________

while groupiter <= 5:  # Iterate 8 times
    title2.append(ws3.cell(row=1, column=int((5*groupiter)-4)).value)  # Create and append title array
    for i in range(1, int((colcount/coldiv)+1)):  # Iterate 5 times
        col = ((groupiter-1)*5) + i  # Move to new column group
    for row in range(3, int(rowcount+1)):  # Go through row for a section, append matrices
        if ws2.cell(row=row, column=col).value is not None:  # Check to ensure there is a value present
            item2.append(ws3.cell(row=row, column=int((5*groupiter)-4)).value)
            CGx2.append(ws3.cell(row=row, column=int((5*groupiter)-3)).value)
            CGy2.append(ws3.cell(row=row, column=int((5*groupiter)-2)).value)
            CGz2.append(ws3.cell(row=row, column=int((5*groupiter)-1)).value)
            Mass2.append(ws3.cell(row=row, column=int(5*groupiter)).value)
        else:
            continue  # If there is no value, move to new group
    groupiter = groupiter + 1

print('MOMENT APPEND END!')

ZipMatrix = zip(item, CGx, CGy, CGz, Mass)  # Combine all into one matrix excluding titles
ZipList = list(ZipMatrix)  # Turn combined matrix into list
ZipList2 = list(zip(item2, CGx2, CGy2, CGz2, Mass2))  # Turn combined moment arm matrix into list

# SMART PART NO. AND NAME LIST POPULATION
sheet4iter = 1
dash = []  # Dash number empty list
SPN = []  # SPN empty list
# Name = []  # Name empty list
sheet4row = ws4.max_row  # Row length of 'Dash Numbers' sheet
sheet4col = ws4.max_column  # Col length of 'Dash Numbers' sheet

for i in range(1, sheet4row+1, 1):  # Iterate through all row entries on sheet 4
    dash.append(ws4.cell(row=i, column=1).value)
    SPN.append(ws4.cell(row=i, column=2).value)
    # Name.append(ws4.cell(row=i, column=3).value)

print('DASH NUMBER RECORDING END!')

DashNumbers = list(zip(dash, SPN))
NF = []
SPNFINAL = []
NAMETEMP = []
SPNTEMP = []
MASKLEN = []

print('Dash Numbers Length:', len(DashNumbers))

# CREATE NAMING MATRIX _________________________________________________________________________________________________________________________________________________________________________________

for i in range(1, len(DashNumbers)+1, 1):  # Iterate through the length of the DashNumbers list
    for j in range(1, len(DashNumbers[i-1][1])+1, 1):  # Iterate through the SPN string
        if 1 <= j <= 6:  # Plane Name and Panel Name
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        elif j == 7:  # Installation Location
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        elif j == 8:  # Oxygen Duration
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        elif j == 9:  # Mask Quantity
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        elif j == 10:  # Streamers
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        elif j == 11:  # Reading Light Quantity
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        elif j == 12:  # Dimmable / Non-Dimmable
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        elif j == 13:  # Wayfinders
            SPNTEMP.append(DashNumbers[i-1][1][j-1])
        else:
            print('    error: No Dash-Number Entry on loop', j)
            print('')
            continue

    # INSTALLATION LOCATION LOOP
    if '1' in SPNTEMP[7-1]:
        NAMETEMP.append('OUTBOARD LH ')  # location
        NAMETEMP.append('39" THROW')  # throw
        MASKLEN.append('55" ')
    elif '2' in SPNTEMP[7-1]:
        NAMETEMP.append('CENTER ')  # location - REG
        NAMETEMP.append('54" THROW')  # throw
        MASKLEN.append('65" ')
    elif '3' in SPNTEMP[7-1]:
        NAMETEMP.append('RESERVED LOCATION ')  # location
        NAMETEMP.append('RESERVED THROW')  # throw
        MASKLEN.append('RESERVED LENGTH')
    elif '4' in SPNTEMP[7-1]:
        NAMETEMP.append('CENTER ')  # location - PREMIUM
        NAMETEMP.append('68" THROW')  # throw
        MASKLEN.append('80" ')
    elif '5' in SPNTEMP[7-1]:
        NAMETEMP.append('OUTBOARD RH ')  # location
        NAMETEMP.append('39" THROW')  # throw
        MASKLEN.append('55" ')
    else:
        print('    error: no installation location')
        print('        conflicting entry:', SPNTEMP[7-1], ' in SPN:', *SPNTEMP)
        print('')

    # OXYGEN DURATION LOOP
    if 'S' in SPNTEMP[8-1]:
        NAMETEMP.append('(SHORT), ')
    elif 'M' in SPNTEMP[8-1]:
        NAMETEMP.append('(MEDIUM), ')
    elif 'L' in SPNTEMP[8-1]:
        NAMETEMP.append('(LONG), ')
    else:
        print('    error: no specified duration')
        print('        conflicting entry:', SPNTEMP[8-1], ' in SPN:', *SPNTEMP)
        print('')

    # MASK QUANTITY
    if '2' in SPNTEMP[9-1]:
        NAMETEMP.append('2 MASK ')
    elif '3' in SPNTEMP[9-1]:
        NAMETEMP.append('3 MASK ')
    elif '4' in SPNTEMP[9-1]:
        NAMETEMP.append('4 MASK ')
    elif '5' in SPNTEMP[9-1]:
        NAMETEMP.append('5 MASK ')
    elif '6' in SPNTEMP[9-1]:
        NAMETEMP.append('6 MASK ')
    else:
        print('    error: no mask quantity')
        print('        conflicting entry:', SPNTEMP[9-1], ' in SPN:', *SPNTEMP)
        print('')

    # STREAMERS
    if 'C' in SPNTEMP[10-1]:
        NAMETEMP.append(', STREAMER')  # 2  streamer
    elif 'D' in SPNTEMP[10-1]:
        NAMETEMP.append(', STREAMER')  # 1 streamer
    elif 'E' in SPNTEMP[10-1]:
        NAMETEMP.append(', STREAMER')  # 2 streamer
    elif 'N' in SPNTEMP[10-1]:
        NAMETEMP.append('')
    else:
        print('    error: no streamer specified')
        print('        conflicting entry:', SPNTEMP[10-1], ' in SPN:', *SPNTEMP)
        print('')

    # QTY READING LIGHTS
    if '2' in SPNTEMP[11-1]:
        NAMETEMP.append('2 LIGHT ')
    elif '3' in SPNTEMP[11-1]:
        NAMETEMP.append('3 LIGHT ')
    elif '4' in SPNTEMP[11-1]:
        NAMETEMP.append('4 LIGHT ')
    else:
        print('    error: no light quantity')
        print('        conflicting entry:', SPNTEMP[11-1], ' in SPN:', *SPNTEMP)
        print('')

    # DIM vs. NON-DIM
    if 'A' in SPNTEMP[12-1]:
        NAMETEMP.append('')
    elif 'B' in SPNTEMP[12-1]:
        NAMETEMP.append('DIMMABLE ')
    else:
        print('    error: no dim/non-dim spec. specified')
        print('        conflicting entry:', SPNTEMP[12-1], ' in SPN:', *SPNTEMP)
        print('')

    # WAYFINDER
    if '0' in SPNTEMP[13-1]:
        NAMETEMP.append('')  # NO WAYFINDERS
    elif '1' in SPNTEMP[13-1]:
        NAMETEMP.append(', WAYFINDER')
    elif '2' in SPNTEMP[13-1]:
        NAMETEMP.append(', WAYFINDER')  # 2 wayfinders
    else:
        print('    error: no wayfinder specified')
        print('        conflicting entry:', SPNTEMP[13-1], ' in SPN:', *SPNTEMP)
        print('')
    NF.append(NAMETEMP)
    SPNFINAL.append(SPNTEMP)
    NAMETEMP = []
    SPNTEMP = []

print('SPN APPEND END!')
print('Length of SPN Table:', len(SPNFINAL), 'entries')

Mx = 0
My = 0
Mz = 0
j = 1
Mxarray = np.zeros(shape=(1, 5))
Myarray = np.zeros(shape=(1, 5))
Mzarray = np.zeros(shape=(1, 5))
Massarray = np.zeros(shape=(1, 5))
coliter = 0
columniter = 1
groupiter = 1
FINAL = []

# ANALYZE AND MATCH SPN TO DESCRIPTION _________________________________________________________________________________________________________________________________________________________________

for i in range(1, len(SPNFINAL)+1):  # Iterates through all SPN's

    TEMPLIST = []
    MxTEMPLIST = []
    MyTEMPLIST = []
    MzTEMPLIST = []
    MaTEMPLIST = []
    FINALPRE = []
    SPN = ''.join(SPNFINAL[i-1])

    ws5.cell(row=i, column=1).value = SPN

    for j in range(1, len(ZipList2)+1):

        # BASE CG GROUP
        if ZipList2[j-1][0][1] == SPNFINAL[i-1][11-1] and ZipList2[j-1][0][2] == SPNFINAL[i-1][12-1] and ZipList2[j-1][0][3] == SPNFINAL[i-1][13-1]:
            MxTEMPLIST.append(ZipList2[j-1][1])
            MyTEMPLIST.append(ZipList2[j-1][2])
            MzTEMPLIST.append(ZipList2[j-1][3])
            MaTEMPLIST.append(ZipList2[j-1][4])

        # CYLINDER GROUP
        if ZipList2[j-1][0][1] == SPNFINAL[i-1][8-1] and ZipList2[j-1][0][2] == SPNFINAL[i-1][9-1]:
            MxTEMPLIST.append(ZipList2[j-1][1])
            MyTEMPLIST.append(ZipList2[j-1][2])
            MzTEMPLIST.append(ZipList2[j-1][3])
            MaTEMPLIST.append(ZipList2[j-1][4])

        # MASK GROUP
        if ZipList2[j-1][0][1] == SPNFINAL[i-1][7-1] and ZipList2[j-1][0][2] == SPNFINAL[i-1][9-1]:
            MxTEMPLIST.append(ZipList2[j-1][1])
            MyTEMPLIST.append(ZipList2[j-1][2])
            MzTEMPLIST.append(ZipList2[j-1][3])
            MaTEMPLIST.append(ZipList2[j-1][4])

        # MASK GROUP 2, RESERVED
        if '3' in SPNFINAL[i-1][7-1] and ZipList2[j-1][0][2] == SPNFINAL[i-1][9-1] and '1' in ZipList2[j-1][0][1]:
            MxTEMPLIST.append(ZipList2[j-1][1])
            MyTEMPLIST.append(ZipList2[j-1][2])
            MzTEMPLIST.append(ZipList2[j-1][3])
            MaTEMPLIST.append(ZipList2[j-1][4])

        # MASK GROUP 3, OUTBOARD RIGHT
        if '5' in SPNFINAL[i-1][7-1] and ZipList2[j-1][0][2] == SPNFINAL[i-1][9-1] and '1' in ZipList2[j-1][0][1]:
            MxTEMPLIST.append(ZipList2[j-1][1])
            MyTEMPLIST.append(ZipList2[j-1][2])
            MzTEMPLIST.append(ZipList2[j-1][3])
            MaTEMPLIST.append(ZipList2[j-1][4])

        # PCBA GROUP
        if ZipList2[j-1][0][1] == SPNFINAL[i-1][9-1] and 'p' in ZipList2[j-1][0][0]:
            MxTEMPLIST.append(ZipList2[j-1][1])
            MyTEMPLIST.append(ZipList2[j-1][2])
            MzTEMPLIST.append(ZipList2[j-1][3])
            MaTEMPLIST.append(ZipList2[j-1][4])

        # STREAMER GROUP
        if ZipList2[j-1][0][1] == SPNFINAL[i-1][10-1] and 's' in ZipList2[j-1][0][0]:
            MxTEMPLIST.append(ZipList2[j-1][1])
            MyTEMPLIST.append(ZipList2[j-1][2])
            MzTEMPLIST.append(ZipList2[j-1][3])
            MaTEMPLIST.append(ZipList2[j-1][4])
        else:
            continue

        MxSum = MxTEMPLIST[0] + MxTEMPLIST[1] + MxTEMPLIST[2] + MxTEMPLIST[3] + MxTEMPLIST[4]
        MySum = MyTEMPLIST[0] + MyTEMPLIST[1] + MyTEMPLIST[2] + MyTEMPLIST[3] + MyTEMPLIST[4]
        MzSum = MzTEMPLIST[0] + MzTEMPLIST[1] + MzTEMPLIST[2] + MzTEMPLIST[3] + MzTEMPLIST[4]
        MaSum = MaTEMPLIST[0] + MaTEMPLIST[1] + MaTEMPLIST[2] + MaTEMPLIST[3] + MaTEMPLIST[4]

        FINALPRE.append(''.join(SPNFINAL[i-1]))
        FINALPRE.append(MxSum)
        FINALPRE.append(MySum)
        FINALPRE.append(MzSum)
        FINALPRE.append(MaSum)
        FINAL.append(FINALPRE)

print('APPEND END!')

l = 0

# print('NF:', len(NF))
# print('Dash: ', len(dash))

# WRITE TO NEW WORKBOOK ________________________________________________________________________________________________________________________________________________________________________________

for row in range(1, len(dash)+1):
    ws5.cell(row=row, column=1).value = dash[row-1]
    M = 'ASM, 777X, CONSOLIDATED PANEL '
    ws5.cell(row=row, column=2).value = ''.join([M, NF[row-1][0],  NF[row-1][3], MASKLEN[row-1], NF[row-1][2], NF[row-1][5], NF[row-1][6], NF[row-1][1], NF[row-1][4], NF[row-1][7]])
    for col in range(3, len(FINAL[row-1])+3):
        ws5.cell(row=row, column=col).value = FINAL[row-1][col-3]
    l = l+1

# print(l, '\n')
print('EXAMPLE NAME OUTPUT:')
print('    ', ws5.cell(row=len(FINAL), column=2).value, '\n')
print('WRITE END!')
wb2.save('BODS (WEIGHT & CGs) CONS DUMP.xlsx')
print('SAVE END!')
